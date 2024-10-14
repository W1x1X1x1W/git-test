from openpyxl import workbook, load_workbook
wb = load_workbook('C:/Users/DELL/Desktop/h.xlsx')
ws = wb.active

ws['A1'].value = 'test_name'






wb.save('excelFile_new.xlsx')