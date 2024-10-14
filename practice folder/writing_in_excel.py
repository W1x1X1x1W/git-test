from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


data = {
	"Joe": {
		"math": 65,
		"science": 78,
		"english": 98,
		"gym": 89
	},
	"Bill": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Tim": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"Sally": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Jane": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}

wb = Workbook()
ws = wb.active
ws.title = 'Grades'

heading1 = ["Name"] + list(data["Jane"].keys())
ws.append(heading1)
for person in data.keys():
    grades = list(data[person].values())
    ws.append([person] + grades)
for col in range(2, len(data['Joe']) + 2):
    char = get_column_letter(col)
    ws[char + "7"] = f"=sum({char + '2'}:{char + '6'})/{len(data)}"

for col1 in range(1, 6):
    ws[get_column_letter(col1) + '1'].font = Font(bold=True, color='FF000000')













wb.save('excelFile_new.xlsx')